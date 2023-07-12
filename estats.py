import openpyxl
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from pruebas_h0vsha import cargar_datos

def filtrar(data, texto):
    columnas_filtradas = [col for col in data.columns if texto in col and 'Otro' not in col]
    datos_filtrados = data[columnas_filtradas]
    return datos_filtrados

def empleo_bgm(data):
    empleo_counts = data.sum()
    empleo_porcentajes = empleo_counts / len(data) * 100
    empleo_porcentajes = empleo_porcentajes.round(2) # Redondeamos
    empleo_tabla = pd.DataFrame({'Tarea': [col.split('(')[-1][:-1] for col in data.columns],
                                 'Total de personas': empleo_counts.values,
                                 'Porcentaje': empleo_porcentajes.values})
    empleo_tabla.name = "Empleo de la música de fondo en las diferentes tareas relacionadas con el estudio"
    return empleo_tabla

def sentimientos(data):
    sent_counts = data.sum()
    sent_porcentajes = sent_counts / len(data) * 100
    sent_porcentajes = sent_porcentajes.round(2) # Redondeamos
    sent_tabla = pd.DataFrame({'Sentimientos/efectos': [col.split('(')[-1][:-1] for col in data.columns],
                                 'Total de personas': sent_counts.values,
                                 'Porcentaje': sent_porcentajes.values})
    sent_tabla.name = "Efecto de la música de fondo sobre el estado anímico de los estudiantes"
    return sent_tabla

def color_degradado(color_base, n_barras):
    r_base, g_base, b_base = int(color_base[1:3], 16), int(color_base[3:5], 16), int(color_base[5:7], 16)
    color_degradado = [
        f'#{hex(int(r_base * (i + 1) / n_barras))[2:]:>02}'
        f'{hex(int(g_base * (i + 1) / n_barras))[2:]:>02}'
        f'{hex(int(b_base * (i + 1) / n_barras))[2:]:>02}'
        for i in range(n_barras)
    ]
    return color_degradado

def bar_d(tabla):
    plt.figure(figsize=(10, 6))
    custom_palette = color_degradado("#840032", len(tabla))  # Color granate degradado
    sns.barplot(x='Total de personas', y=str(tabla.columns[0]), data=tabla, palette=custom_palette)
    
    for i, (_, row) in enumerate(tabla.iterrows()):
        count = row['Total de personas']
        percent = row['Porcentaje']
        if i == len(tabla) - 1:  # Ajustar posición del porcentaje en la última barra
            plt.text(count + 0.5, i, f'{percent:.1f}%', va='center', ha='left')
        else:
            plt.text(count + 1, i, f'{percent:.1f}%', va='center')

    plt.title(tabla.name, fontweight='bold', pad=20)  # Título en negrita y con mayor separación
    plt.xlabel('Total de personas')
    plt.ylabel(str(tabla.columns[0]))
    plt.xlim(0, 150)  # Ajustar límites del eje x
    plt.tight_layout()
    # plt.show()

    ruta_guardar = 'F:/Cursos y master/Master educación/TFM/Resultados/Graficos de barras/'
    nombre_archivo = tabla.name.split()[0] + '_bgm' +'.png'
    plt.savefig(ruta_guardar + nombre_archivo)
    print("\nEl gráfico se ha guardado exitosamente.")

def guardar_tabla_excel(tabla):
    ruta_archivo_excel = 'F:/Cursos y master/Master educación/TFM/Resultados/Resultados.xlsx'
    libro_excel = openpyxl.load_workbook(ruta_archivo_excel)
    nombre_hoja = tabla.name.split()[0]  # Obtener la primera palabra del título de la tabla

    # Crear una nueva hoja en el libro de Excel
    nueva_hoja = libro_excel.create_sheet(title=nombre_hoja)

    # Escribir los encabezados de columna en la nueva hoja
    for j, columna in enumerate(tabla.columns, start=1):
        nueva_hoja.cell(row=1, column=j, value=columna)

    # Escribir los datos de la tabla en la nueva hoja
    for i, (_, row) in enumerate(tabla.iterrows(), start=2):
        for j, value in enumerate(row, start=1):
            nueva_hoja.cell(row=i, column=j, value=value)

    # Guardar el libro de Excel con la nueva hoja
    libro_excel.save(ruta_archivo_excel)
    print(f"La tabla se ha guardado exitosamente en la hoja '{nombre_hoja}' del archivo '{ruta_archivo_excel}'.")

def main():
    # Cargar los datos
    datos = cargar_datos()
    # Filtrar los datos
    datos_filtrados = filtrar(datos, 'Q10')
    # Contar empleo_bgm
    empleo_tabla = empleo_bgm(datos_filtrados)
    # Mostrar la tabla
    print(empleo_tabla.to_string(index=False))
    # Crear el diagrama de barras y guardar
    bar_d(empleo_tabla)
    # Sentimientos
    sent_tabla = sentimientos (filtrar(datos, 'Q21'))
    # Mostrar la tabla
    print(sent_tabla.to_string(index=False))
    # Crear el diagrama de barras y guardar
    bar_d(sent_tabla)
    # Guardar en excel
    guardar_tabla_excel(empleo_tabla)
    guardar_tabla_excel(sent_tabla)

if __name__ == "__main__":
    main()

