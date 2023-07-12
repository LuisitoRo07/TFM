import textwrap
import openpyxl
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from pruebas_h0vsha import cargar_datos

def filtrar(data, texto):
    columnas_filtradas = [col for col in data.columns if texto in col and 'Otro' not in col]
    datos_filtrados = data[columnas_filtradas]
    return datos_filtrados

def tratar(data):
    mapeo = {1: 'Nunca', 2: 'Muy Poco', 3: 'Poco', 4: 'Algo', 5: 'Bastante', 6: 'Mucho'}
    data_tratada = data.replace(mapeo)
    return data_tratada

def est_columna(data, n_col):
    # Obtener el nombre de la columna para usarlo como título
    nombre_columna = str(data.columns[n_col])
    # Extraer el texto entre paréntesis y convertir la primera letra en mayúscula
    titulo = nombre_columna[nombre_columna.find("(") + 1:nombre_columna.find(")")].capitalize()

    # Ordenar los niveles de identificación en el orden deseado
    niveles_ordenados = ['Nunca', 'Muy Poco', 'Poco', 'Algo', 'Bastante', 'Mucho']
    data[nombre_columna] = pd.Categorical(data[nombre_columna], categories=niveles_ordenados, ordered=True)
    
    # Contar el total en cada grupo
    conteo_grupos = data.groupby(nombre_columna).size()
    total_alumnos = conteo_grupos.sum()

    # Calcular el porcentaje de cada grupo
    porcentaje_grupos = conteo_grupos / total_alumnos * 100

    # Crear la tabla de resultados
    tabla_resultados = pd.DataFrame({
        'Nivel de identificación': conteo_grupos.index,
        'Alumnos': conteo_grupos.values,
        'Porcentaje': porcentaje_grupos.values
    })

    # Establecer el título de la tabla
    tabla_resultados.name = titulo
    # Mostrar por pantalla
    print('\n')
    print(titulo)
    print('__________________________________________________________________________________________________________________________')
    print(tabla_resultados)
    # Devolver a main
    return tabla_resultados

def color_degradado(color_base, n_barras):
    r_base, g_base, b_base = int(color_base[1:3], 16), int(color_base[3:5], 16), int(color_base[5:7], 16)
    color_degradado = [
        f'#{hex(int(r_base * (i + 1) / n_barras))[2:]:>02}'
        f'{hex(int(g_base * (i + 1) / n_barras))[2:]:>02}'
        f'{hex(int(b_base * (i + 1) / n_barras))[2:]:>02}'
        for i in range(n_barras)
    ]
    return color_degradado

def bar_d(tabla, n_col):
    plt.figure(figsize=(10, 6))
    custom_palette = color_degradado("#840032", len(tabla))  # Color granate degradado
    sns.barplot(x='Alumnos', y=str(tabla.columns[0]), data=tabla, palette=custom_palette)
    
    for i, (_, row) in enumerate(tabla.iterrows()):
        count = row['Alumnos']
        percent = row['Porcentaje']
        if i == len(tabla) - 1:  # Ajustar posición del porcentaje en la última barra
            plt.text(count + 0.5, i, f'{percent:.1f}%', va='center', ha='left')
        else:
            plt.text(count + 1, i, f'{percent:.1f}%', va='center')

    # Ajustar título a varias líneas si es necesario
    titulo = tabla.name
    titulo_wrapped = "\n".join(textwrap.wrap(titulo, width=30))

    plt.title(f'"{titulo_wrapped}"', fontweight='bold', style='italic', pad=20)  # Título entre comillas y en cursiva
    plt.xlabel('Alumnos')
    plt.ylabel(str(tabla.columns[0]))
    plt.xlim(0, 100)  # Ajustar límites del eje x
    plt.tight_layout()
    #plt.show()

    ruta_guardar = 'F:/Cursos y master/Master educación/TFM/Resultados/Graficos de barras/Cognitivo/'
    nombre_archivo = 'Afirmacion' + str(n_col+1) +'.png'
    plt.savefig(ruta_guardar + nombre_archivo)
    print("\nEl gráfico se ha guardado exitosamente.\n**************************************************")

def guardar_tabla_excel(tabla, n_col):
    ruta_archivo_excel = 'F:/Cursos y master/Master educación/TFM/Resultados/Resultados_cognitivo.xlsx'
    
    try:
        # Intentar cargar el libro de Excel existente
        libro_excel = openpyxl.load_workbook(ruta_archivo_excel)
    except FileNotFoundError:
        # Si el archivo no existe, crear un nuevo libro de Excel
        libro_excel = openpyxl.Workbook()
    
    nombre_hoja = 'Afirmacion ' + str(n_col+1)  

    # Crear una nueva hoja en el libro de Excel
    nueva_hoja = libro_excel.create_sheet(title=nombre_hoja)

    # Escribir el título de la tabla en la nueva hoja
    nueva_hoja.cell(row=1, column=1, value=tabla.name)

    # Escribir los encabezados de columna en la nueva hoja
    for j, columna in enumerate(tabla.columns, start=1):
        nueva_hoja.cell(row=2, column=j, value=columna)

    # Escribir los datos de la tabla en la nueva hoja
    for i, (_, row) in enumerate(tabla.iterrows(), start=3):
        for j, value in enumerate(row, start=1):
            nueva_hoja.cell(row=i, column=j, value=value)

    # Guardar el libro de Excel con la nueva hoja
    libro_excel.save(ruta_archivo_excel)
    print(f"\nLa tabla se ha guardado exitosamente en la hoja '{nombre_hoja}' del archivo '{ruta_archivo_excel}'.")

def est_y_graf (datos_tratados, n_col):
    # Obtener estadísticas de la columna seleccionada
    tabla_estadisticas = est_columna(datos_tratados, n_col)
    #Guardar tabla a excel
    guardar_tabla_excel(tabla_estadisticas, n_col)
    # Hacer gráfico 
    bar_d (tabla_estadisticas, n_col)
    
    index_last_col = datos_tratados.shape[1] - 1
    if n_col < index_last_col:
        n_col += 1
        est_y_graf (datos_tratados, n_col)


def main():
    # Cargar los datos
    datos = cargar_datos()
    # Filtrar los datos
    datos_filtrados = filtrar(datos, 'Q22')
    # Tratar los datos filtrados
    datos_tratados = tratar(datos_filtrados)
    # Hacer la estadística y gráficos de cada columna
    est_y_graf (datos_tratados, 0) # Inicializamos con n_col = 0
    

if __name__ == "__main__":
    main()

