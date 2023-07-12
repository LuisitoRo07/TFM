import sys
from matplotlib.lines import Line2D
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import pingouin as pg
from statsmodels.graphics.factorplots import interaction_plot
from scipy.stats import kruskal

def cargar_datos ():
    # Ruta del archivo Excel
    ruta_excel = 'F:/Cursos y master/Master educación/TFM/Resultados/Muestra_localidadVLL.xlsx'
    # Leer los datos del archivo Excel
    datos = pd.read_excel(ruta_excel)
    return datos

def elegir_estudio(datos, confirmacion):
    # Mostrar el total de datos y solicitar confirmación
    N_question = "Q8" #La cuestión que quiero buscar (columna), en este caso la de la frecuencia de estudio con BGM
    #print("Elige la asignatura de estudio:\n Q - Química  D - Dibujo\n FQ - Fisca-Química  QF - Química-Física\n MF - Mates-Fis  MQ - Mates-Qui")
    #confirmacion = input("Antes del guión es frecuencia de BGM, después es nota: ")
    #ahora elegimos las columnas (en cuanto a notas, F-69  Q-70  D-71)
    if confirmacion == "F":
        subject = "Física"
        c_nota = 69
        subject2 = subject
        sheet_name = "frec_Fisica"
    elif confirmacion == "Q":
        subject = "Química"
        c_nota = 70
        subject2 = subject
        sheet_name = "frec_Quimica"
    elif confirmacion == "D":
        subject = "Dibujo Técnico"
        c_nota = 71
        subject2 = subject
        sheet_name = "frec_Dibujo"
    elif confirmacion == "FQ":
        subject = "Física"
        c_nota = 70
        subject2 = 'Química'
        sheet_name = "frec_FisQui"
    elif confirmacion == "QF":
        subject = "Química"
        c_nota = 69
        subject2 = 'Física'
        sheet_name = "frec_QuiFis"
    elif confirmacion == "MF":
        subject = "Matemáticas"
        c_nota = 69
        subject2 = 'Física'
        sheet_name = "frec_MateFis"
    elif confirmacion == "MQ":
        subject = "Matemáticas"
        c_nota = 70
        subject2 = 'Química'
        sheet_name = "frec_MateQui"
    else:
        sys.exit('El dato introducido no era una opción')
        
    #Seleccionamos el índice de la columna frec_subject
    f_colum = np.where(datos.columns.str.contains(f'{N_question}.*{subject}'))[0][0]
    # Seleccionar las columnas necesarias
    datos_def = datos.iloc[:, [0, f_colum, c_nota]]
    # Eliminar filas con valores en blanco o no numéricos
    datos_def = datos_def.dropna()
    datos_def.iloc[:, 1] = pd.to_numeric(datos_def.iloc[:, 1], errors='coerce')
    # Mapear los valores de la columna de frecuencia de estudio
    mapeo_frecuencia = {1: 'Nunca', 2: 'Muy Poco', 3: 'Poco', 4: 'Algo', 5: 'Bastante', 6: 'Mucho'}
    datos_def.iloc[:, 1] = datos_def.iloc[:, 1].map(mapeo_frecuencia)
    # Cambiar los títulos de las columnas
    nombre_last_c = 'Nota Media ' + subject2
    columna_frec = 'Frecuencia BGM ' + subject
    nombres_columnas = ['ID', columna_frec, nombre_last_c]
    datos_def.columns = nombres_columnas
    # Ordenar los datos por la columna columna_frec en el orden especificado
    orden_frecuencia = ['Nunca', 'Muy Poco', 'Poco', 'Algo', 'Bastante', 'Mucho']
    datos_def = datos_def.loc[datos_def[columna_frec].isin(orden_frecuencia)]
    datos_def[columna_frec] = pd.Categorical(datos_def[columna_frec], categories=orden_frecuencia, ordered=True)
    datos_def = datos_def.sort_values(columna_frec)
    
    return datos_def, sheet_name, nombre_last_c, subject2, columna_frec, subject

def save_excel(datos1, datos2, datos3, datos4, datos5, datos6, sheet):
    # Ruta del archivo Excel existente
    ruta_archivo_excel = 'F:/Cursos y master/Master educación/TFM/Resultados/Resultados.xlsx'
    # Leer el archivo Excel existente
    book = load_workbook(ruta_archivo_excel)
    # Crear un objeto de escritura Excel
    writer = pd.ExcelWriter(ruta_archivo_excel, engine='openpyxl')
    writer.book = book
    # Obtener la hoja existente o crear una nueva
    if sheet in book.sheetnames:
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        hoja_existente = writer.sheets[sheet]
        startrow = hoja_existente.max_row + 3  # Obtener la última fila y agregar 3 filas de separación
    else:
        hoja_existente = book.create_sheet(sheet)
        startrow = 1  # Comenzar en la primera fila si la hoja es nueva
    # Guardar la tabla 1 en la hoja
    hoja_existente.cell(row=startrow, column=1, value='Resumen estadístico de cada grupo')
    startrow += 1  # Actualizar la fila de inicio para la siguiente tabla
    datos1.to_excel(writer, sheet_name=sheet, startrow=startrow, index=True)
    startrow += datos1.shape[0] + 3  # Actualizar la fila de inicio para la siguiente tabla
    # Guardar la tabla 2 en la hoja
    hoja_existente.cell(row=startrow, column=1, value='Shapiro-Wilk normality test')
    startrow += 1  # Actualizar la fila de inicio para la siguiente tabla
    datos2.to_excel(writer, sheet_name=sheet, startrow=startrow, index=False)
    startrow += datos2.shape[0] + 3  # Actualizar la fila de inicio para la siguiente tabla    
    # Guardar la tabla 3 en la hoja
    hoja_existente.cell(row=startrow, column=1, value='Levene homoscedasticity test')
    startrow += 1  # Actualizar la fila de inicio para la siguiente tabla
    datos3.to_excel(writer, sheet_name=sheet, startrow=startrow, index=False)
    startrow += datos3.shape[0] + 3  # Actualizar la fila de inicio para la siguiente tabla
    # Guardar la tabla 4 en la hoja
    hoja_existente.cell(row=startrow, column=1, value='ANOVA')
    startrow += 1  # Actualizar la fila de inicio para la siguiente tabla
    datos4.to_excel(writer, sheet_name=sheet, startrow=startrow, index=False)
    startrow += datos4.shape[0] + 3  # Actualizar la fila de inicio para la siguiente tabla
    # Guardar la tabla 5 en la hoja
    hoja_existente.cell(row=startrow, column=1, value='Kruskal Wallis no parametric test')
    startrow += 1  # Actualizar la fila de inicio para la siguiente tabla
    datos5.to_excel(writer, sheet_name=sheet, startrow=startrow, index=False)
    startrow += datos5.shape[0] + 3  # Actualizar la fila de inicio para la siguiente tabla
    # Guardar la tabla 6 en la hoja
    hoja_existente.cell(row=startrow, column=1, value='Turkey\'s HDS test')
    startrow += 1  # Actualizar la fila de inicio para la siguiente tabla
    datos6.to_excel(writer, sheet_name=sheet, startrow=startrow, index=False)
    # Cerrar el objeto de escritura Excel
    writer.save()
    # Mostrar el mensaje de confirmación
    print(f"La hoja '{sheet}' se ha añadido al archivo Excel: {ruta_archivo_excel}")
  
def estadísticas_iniciales (datos, columna_media, columna_frec):
    # Contar el total en cada grupo
    conteo_grupos = datos.groupby(columna_frec).size()
    # Calcular la media y la desviación estándar de la última columna
    media = datos.groupby(columna_frec)[columna_media].mean()
    desviacion_tipica = datos.groupby(columna_frec)[columna_media].std()
    # Crear una tabla con las medidas estadísticas
    medidas_estadisticas = pd.DataFrame({'Media': media, 'Desviación Típica': desviacion_tipica})
    # Unir el conteo en cada grupo y las medidas estadísticas en una misma tabla
    tabla_estadisticas = conteo_grupos.to_frame().join(medidas_estadisticas)
    # Cambiar el título de la columna
    tabla_estadisticas = tabla_estadisticas.rename(columns={0: 'Total de alumnos'})
    # Imprimir la tabla de medidas estadísticas
    # print(tabla_estadisticas)
    # print("\n")
    return tabla_estadisticas

def box_diagram (datos, columna_media, columna_frec, subject, subject2):
    # # Definir una paleta de colores personalizada
    colores = ['#FF7F50', '#1E90FF', '#FFD700', '#32CD32', '#FF69B4', '#00CED1']

    # Establecer estilo de los gráficos y la paleta de colores
    sns.set(style="whitegrid", palette=colores)

    # Representar los grupos en diagramas de caja
    plt.figure(figsize=(8, 6))
    ax = sns.boxplot(x=columna_frec, y=columna_media, data=datos, showmeans=True, meanline=True, meanprops=dict(color="black", linewidth=1.5))

    # Personalizar los ejes
    plt.xlabel(columna_frec, fontweight='bold', labelpad=10)
    plt.ylabel(columna_media, fontweight='bold')
    plt.title('Comparación de notas medias en función del uso de BGM', fontweight='bold')

    # Añadir leyenda
    legend_labels = ['Media']
    legend_lines = [Line2D([0], [0], color='black', linestyle='--', linewidth=1.5)]
    ax.legend(legend_lines, legend_labels)

    # Ajustar el espaciado de los ejes
    plt.xticks(fontweight='bold')
    plt.yticks(fontweight='bold')

    # Guardar el diagrama como imagen en F:\Cursos y master\Master educación\TFM\Resultados\Diagramas de cajas
    nombre_imagen = f"F:/Cursos y master/Master educación/TFM/Resultados/Diagramas de cajas/Box_diag_{subject}-{subject2}.png"
    plt.tight_layout()
    plt.savefig(nombre_imagen)
    plt.close()

    # Mostrar el mensaje de confirmación
    print(f"El diagrama de caja se ha guardado como imagen: {nombre_imagen}")

def tests_previos (datos, columna_media, columna_frec):
    # Obtener los grupos únicos de Frecuencia BGM
    grupos = datos[columna_frec].unique()

    # Crear una lista para almacenar los resultados
    resultados = []

    # Realizar el test de normalidad Shapiro-Wilk para cada grupo
    for grupo in grupos:
        datos_grupo = datos[datos[columna_frec] == grupo]
        resultado_test = pg.normality(datos_grupo[columna_media], method='shapiro')
        resultados.append({
            'Grupo': grupo,
            'W': resultado_test['W'].values[0],
            'p-value': resultado_test['pval'].values[0],
            'Normalidad': resultado_test['normal'].values[0]
        })

    # Crear un DataFrame con los resultados
    tabla_resultados = pd.DataFrame(resultados)

    # Imprimir la tabla de resultados
    # print(tabla_resultados)
    # print('\n')
    
    # Realizar el test de homocedasticidad (Levene)
    levene_test = pg.homoscedasticity(data=datos, dv=columna_media, group=columna_frec)
    levene_results = levene_test[['W', 'pval', 'equal_var']]
    levene_results.columns = ['Estadístico de Levene', 'Valor p', 'Igualdad de varianzas']

    # print("Resultado del test de homocedasticidad (Levene):")
    # print(levene_test)   
    
    return tabla_resultados, levene_results

def anova (datos, columna_media, columna_frec):
    #Anova de una vía
    anova_result = pg.anova(data=datos, dv=columna_media, between=columna_frec, detailed=True)
    #mostramos anova
    # print(anova_result)
    # print("\n")
    return anova_result

def no_parametric_test (datos, columna_media, columna_frec):
    # Realizar el test de Kruskal-Wallis
    kruskal_result = pg.kruskal(data=datos, dv=columna_media, between=columna_frec)

    # Imprimir el resultado detallado
    # print("Resultado detallado del test de Kruskal-Wallis:")
    # print(kruskal_result)
    # print("\n")
    
    return kruskal_result

def hsdturkey_test (datos, columna_media, columna_frec):
    # Realizar el análisis post hoc con el test Tukey HSD
    posthoc_result = pg.pairwise_tukey(data=datos, dv=columna_media, between=columna_frec)

    # Mostrar los resultados post hoc
    # print("Resultados del test Tukey HSD:")
    # print(posthoc_result)
    # print("\n")
    
    return posthoc_result

def main(estudio):
    datos_0 = cargar_datos ()    
    datos, sheet_name, columna_media, subject2, columna_frec, subject = elegir_estudio(datos_0, estudio)
    tabla_estats_0 = estadísticas_iniciales(datos, columna_media, columna_frec)
    #box_diagram(datos,columna_media,columna_frec, subject, subject2)
    shapiro_wilk_test, levene_test = tests_previos (datos, columna_media, columna_frec)
    anova_test = anova (datos, columna_media, columna_frec)
    kruskal_wallis_test = no_parametric_test (datos, columna_media, columna_frec)
    turkey_test = hsdturkey_test (datos, columna_media, columna_frec)
    save_excel (tabla_estats_0, shapiro_wilk_test, levene_test, anova_test, kruskal_wallis_test, turkey_test, sheet_name)    
    
if __name__ == "__main__":
    #Definimos la lista para iterar entre las opciones de estudio
    estudios = ['F','Q','D','FQ','QF','MF','MQ']
    #ahora iteramos para cada estudio y listo
    for estudio in estudios:
        main (str(estudio))