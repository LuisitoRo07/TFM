import pandas as pd
from openpyxl import Workbook, load_workbook

def to_excel(rute, sheet_name, *tablas):
    # Leer el archivo Excel existente o crear uno nuevo si no existe
    try:
        book = load_workbook(rute)
    except FileNotFoundError:
        book = Workbook()
    
    # Crear un objeto de escritura Excel
    writer = pd.ExcelWriter(rute, engine='openpyxl')
    writer.book = book
    
    # Obtener la hoja existente o crear una nueva
    if sheet_name in book.sheetnames:
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        hoja_existente = writer.sheets[sheet_name]
        startrow = hoja_existente.max_row + 3  # Obtener la última fila y agregar 3 filas de separación
    else:
        hoja_existente = book.create_sheet(sheet_name)
        startrow = 1  # Comenzar en la primera fila si la hoja es nueva
    
    # Guardar las tablas en la hoja
    for tabla in tablas:
        #hoja_existente.cell(row=startrow, column=1, value=tabla['title'])
        startrow += 1  # Actualizar la fila de inicio para la siguiente tabla
        df_tabla = tabla
        df_tabla.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
        startrow += df_tabla.shape[0] + 3  # Actualizar la fila de inicio para la siguiente tabla
    
    # Cerrar el objeto de escritura Excel
    writer.save()
    
    return 1
